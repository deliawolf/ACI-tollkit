#!/usr/bin/env python3

import requests
import json
import urllib3
from urllib3.exceptions import InsecureRequestWarning
import xml.etree.ElementTree as ET
import xml.dom.minidom
from tabulate import tabulate
from datetime import datetime
import sys
import getpass
import re
import argparse
import time
import os
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed

# Load environment variables
load_dotenv()
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Disable SSL warnings
urllib3.disable_warnings(InsecureRequestWarning)

class ACIInterfaceInfo:
    def __init__(self, apic_url, username, password):
        """Initialize connection to APIC
        Args:
            apic_url (str): APIC URL (e.g., https://apic)
            username (str): APIC username
            password (str): APIC password
        """
        self.apic_url = apic_url.rstrip('/')  # Remove trailing slash if present
        self.username = username
        self.password = password
        self.session = requests.Session()
        self.session.verify = False
        self.refresh_timeout = None
        self.token = None
        self.interfaces = []
        self.debug = False  # Initialize debug flag
        
    def set_debug(self, enabled=True):
        """Enable or disable debug output"""
        self.debug = enabled
    
    def debug_print(self, *args, **kwargs):
        """Print only if debug is enabled"""
        if self.debug:
            print(*args, **kwargs)

    def _refresh_token(self):
        """Refresh the authentication token before timeout"""
        if not self.token or not self.refresh_timeout:
            return self.login()
            
        current_time = time.time()
        if current_time >= self.refresh_timeout - 60:  # Refresh 60 seconds before timeout
            self.debug_print("Token about to expire, refreshing...")
            return self.login()
        return True

    def list_auth_domains(self):
        """Get available authentication domains from APIC"""
        domains_url = f"{self.apic_url}/api/aaaListDomains.json"
        try:
            print(f"\nGetting authentication domains from: {domains_url}")
            response = self.session.get(
                domains_url,
                timeout=10
            )
            
            print(f"Response status code: {response.status_code}")
            if response.status_code == 200:
                data = response.json()
                domains = []
                for domain in data.get('imdata', []):
                    if isinstance(domain, dict):
                        # Each domain has name and type directly in the imdata array
                        name = domain.get('name', '')
                        domain_type = domain.get('type', '')
                        if name:
                            domains.append({'name': name, 'type': domain_type})
                return domains
            else:
                print(f"Error getting domains. Status code: {response.status_code}")
                print(f"Error Content: {response.text}")
                return None
                
        except requests.exceptions.RequestException as e:
            print(f"Failed to get authentication domains: {str(e)}")
            return None
        except Exception as e:
            print(f"Unexpected error while getting domains: {str(e)}")
            return None

    def login(self, selected_domain="DefaultAuth"):
        """Login to APIC and get token"""
        try:
            # Format username with domain
            if selected_domain == "DefaultAuth":
                domain_username = self.username  # Don't add domain prefix for DefaultAuth
            else:
                domain_username = f"apic:{selected_domain}\\{self.username}"
            print(f"\nLogging in with username: {domain_username}")
            
            login_url = f"{self.apic_url}/api/aaaLogin.json"
            payload = {
                "aaaUser": {
                    "attributes": {
                        "name": domain_username,
                        "pwd": self.password
                    }
                }
            }
            
            headers = {
                'Content-Type': 'application/json'
            }
            
            response = self.session.post(
                login_url,
                headers=headers,
                json=payload,
                timeout=10
            )
            
            if response.status_code == 200:
                data = response.json()
                token = data['imdata'][0]['aaaLogin']['attributes']['token']
                timeout = int(data['imdata'][0]['aaaLogin']['attributes'].get('refreshTimeoutSeconds', '600'))
                
                self.token = token
                self.refresh_timeout = time.time() + timeout
                self.session.headers.update({
                    'APIC-Cookie': token
                })
                print("Successfully logged in to APIC")
                return True
            else:
                if response.text:
                    error_data = response.json()
                    error_message = error_data.get('imdata', [{}])[0].get('error', {}).get('attributes', {}).get('text', 'Unknown error')
                    print(f"Login failed: {error_message}")
                else:
                    print(f"Login failed with status code: {response.status_code}")
                return False
                
        except requests.exceptions.RequestException as e:
            print(f"Network error during login: {str(e)}")
            return False
        except Exception as e:
            print(f"Unexpected error during login: {str(e)}")
            return False

    def _make_request(self, method, url, **kwargs):
        """Make a request to APIC with automatic token refresh"""
        # Try to refresh token if needed
        if not self._refresh_token():
            return None
            
        try:
            response = self.session.request(method, url, **kwargs)
            if response.status_code == 403:  # Token might have expired
                self.debug_print("Got 403, trying to refresh token...")
                if self._refresh_token():  # Try refresh and retry request
                    response = self.session.request(method, url, **kwargs)
            return response
        except Exception as e:
            print(f"Error making request: {str(e)}")
            return None

    def get_interface_info(self):
        """Get interface information using l1PhysIf class and combine with hostname info"""
        interface_data = self._get_raw_interface_info()
        hostname_data = self.get_hostname_info()
        return self.combine_interface_hostname_info(interface_data, hostname_data)

    def _get_raw_interface_info(self):
        """Get raw interface information using l1PhysIf class"""
        query_url = f"{self.apic_url}/api/node/class/l1PhysIf.json"
        query_params = {
            'order-by': 'l1PhysIf.modTs|desc'
        }
        
        response = self._make_request('GET', query_url, params=query_params)
        if response and response.status_code == 200:
            return response.json()
        else:
            self.debug_print(f"Error Content: {response.text if response else 'No response'}")
            return None

    def get_hostname_info(self):
        """Get hostname information using fabricNode class"""
        query_url = f"{self.apic_url}/api/node/class/fabricNode.json"
        query_params = {
            'order-by': 'fabricNode.modTs|desc'
        }
        
        response = self._make_request('GET', query_url, params=query_params)
        if response and response.status_code == 200:
            return response.json()
        else:
            self.debug_print(f"Error Content: {response.text if response else 'No response'}")
            return None

    def combine_interface_hostname_info(self, interface_data, hostname_data):
        """Combine interface and hostname information"""
        if not interface_data or not hostname_data:
            return []

        # Create a mapping of node IDs to hostname info
        node_info = {}
        for item in hostname_data.get('imdata', []):
            if 'fabricNode' in item:
                attrs = item['fabricNode']['attributes']
                node_id = attrs.get('id')
                if node_id:
                    node_info[node_id] = {
                        'name': attrs.get('name', ''),
                        'model': attrs.get('model', ''),
                        'role': attrs.get('role', '')
                    }

        # Combine interface data with hostname info
        combined_data = []
        for item in interface_data.get('imdata', []):
            if 'l1PhysIf' in item:
                interface = item['l1PhysIf']['attributes']
                dn = interface.get('dn', '')
                # Extract node ID from DN (format: topology/pod-X/node-Y/sys/...)
                node_match = re.search(r'node-(\d+)', dn)
                if node_match:
                    node_id = node_match.group(1)
                    host_info = node_info.get(node_id, {})
                    interface.update({
                        'hostname': host_info.get('name', ''),
                        'nodeModel': host_info.get('model', ''),
                        'nodeRole': host_info.get('role', '')
                    })
                combined_data.append(interface)

        return combined_data

    def get_transceiver_info(self):
        """Get transceiver information using ethpmFcot class"""
        query_url = f"{self.apic_url}/api/node/class/ethpmFcot.json"
        query_params = {
            'order-by': 'ethpmFcot.modTs|desc'
        }
        
        response = self._make_request('GET', query_url, params=query_params)
        if response and response.status_code == 200:
            return response.json()
        else:
            self.debug_print(f"Error Content: {response.text if response else 'No response'}")
            return None

    def get_interface_faults(self):
        """Get interface faults"""
        query_url = f"{self.apic_url}/api/node/class/faultInst.json?query-target-filter=and(wcard(faultInst.cause,\"interface\"))"
        query_params = {
            'order-by': 'faultInst.modTs|desc'
        }
        
        response = self._make_request('GET', query_url, params=query_params)
        if response and response.status_code == 200:
            return response.json()
        else:
            self.debug_print(f"Error Content: {response.text if response else 'No response'}")
            return None

    def get_ethernet_statistics(self, interface_dn):
        """Get Ethernet statistics for a specific interface using dbgEtherStats endpoint"""
        # Convert interface DN to Ethernet statistics DN format
        # Example: topology/pod-1/node-123/sys/phys-[eth1/1] -> topology/pod-1/node-123/sys/phys-[eth1/1]/dbgEtherStats
        etherstats_dn = f"{interface_dn}/dbgEtherStats"
        query_url = f"{self.apic_url}/api/node/mo/{etherstats_dn}.json?query-target=self"
        
        try:
            response = self._make_request('GET', query_url)
            if response and response.status_code == 200:
                data = response.json()
                if data.get('imdata') and len(data['imdata']) > 0:
                    etherstats = data['imdata'][0].get('rmonEtherStats', {}).get('attributes', {})
                    return {
                        'broadcastPkts': etherstats.get('broadcastPkts', '0'),
                        'clearTs': etherstats.get('clearTs', 'never'),
                        'multicastPkts': etherstats.get('multicastPkts', '0'),
                        'cRCAlignErrors': etherstats.get('cRCAlignErrors', '0'),
                        'dropEvents': etherstats.get('dropEvents', '0'),
                        'fragments': etherstats.get('fragments', '0'),
                        'jabbers': etherstats.get('jabbers', '0'),
                        'pkts': etherstats.get('pkts', '0'),
                        'oversizePkts': etherstats.get('oversizePkts', '0'),
                        'rXNoErrors': etherstats.get('rXNoErrors', '0'),
                        'rxGiantPkts': etherstats.get('rxGiantPkts', '0'),
                        'rxOversizePkts': etherstats.get('rxOversizePkts', '0'),
                        'tXNoErrors': etherstats.get('tXNoErrors', '0'),
                        'txGiantPkts': etherstats.get('txGiantPkts', '0'),
                        'txOversizePkts': etherstats.get('txOversizePkts', '0'),
                        'undersizePkts': etherstats.get('undersizePkts', '0')
                    }
                else:
                    self.debug_print(f"No Ethernet statistics data found for {etherstats_dn}")
                    return self._get_default_etherstats()
            else:
                self.debug_print(f"Error getting Ethernet statistics for {etherstats_dn}: {response.text if response else 'No response'}")
                return self._get_default_etherstats()
        except Exception as e:
            self.debug_print(f"Exception getting Ethernet statistics for {etherstats_dn}: {str(e)}")
            return self._get_default_etherstats()

    def _get_default_etherstats(self):
        """Return default Ethernet statistics when data is not available"""
        return {
            'broadcastPkts': '0',
            'clearTs': 'never',
            'multicastPkts': '0',
            'cRCAlignErrors': '0',
            'dropEvents': '0',
            'fragments': '0',
            'jabbers': '0',
            'pkts': '0',
            'oversizePkts': '0',
            'rXNoErrors': '0',
            'rxGiantPkts': '0',
            'rxOversizePkts': '0',
            'tXNoErrors': '0',
            'txGiantPkts': '0',
            'txOversizePkts': '0',
            'undersizePkts': '0'
        }

    def get_transceiver_dn(self, interface_dn):
        """Convert interface DN to transceiver DN format"""
        # Example interface DN: topology/pod-1/node-101/sys/phys-[eth1/1]
        # Example transceiver DN: topology/pod-1/node-101/sys/phys-[eth1/1]/phys/fcot
        return f"{interface_dn}/phys/fcot"

    def parse_dn(self, dn):
        """Parse DN string to extract pod, node, and interface
        Examples:
        - Ethernet: topology/pod-1/node-101/sys/phys-[eth1/1]
        - Port-channel: topology/pod-1/node-101/sys/aggr-[po10]
        - Loopback: topology/pod-1/node-101/sys/lb-[lo10]
        - FEX: topology/pod-1/node-101/sys/phys-[eth101/1/1]
        """
        pod = node = interface = None
        
        # Extract pod number
        pod_match = re.search(r'pod-(\d+)', dn)
        if pod_match:
            pod = pod_match.group(1)
            
        # Extract node number
        node_match = re.search(r'node-(\d+)', dn)
        if node_match:
            node = node_match.group(1)
            
        # Extract interface name
        # Try ethernet interface first (including FEX)
        eth_match = re.search(r'phys-\[(?:eth)?(\d+(?:/\d+){1,2})\]', dn)
        if eth_match:
            interface = f"eth{eth_match.group(1)}"
        else:
            # Try port-channel
            po_match = re.search(r'aggr-\[(?:po)?(\d+)\]', dn)
            if po_match:
                interface = f"po{po_match.group(1)}"
            else:
                # Try loopback
                lo_match = re.search(r'lb-\[(?:lo)?(\d+)\]', dn)
                if lo_match:
                    interface = f"lo{lo_match.group(1)}"
                
        self.debug_print(f"Parsed DN: pod={pod}, node={node}, interface={interface}")
        return pod, node, interface

    def get_physical_details(self, dn):
        """Get physical interface details including operational state and reason."""
        # Example input DN: topology/pod-1/node-101/sys/phys-[eth1/14]
        # We need: topology/pod-1/node-101/sys/phys-[eth1/14]/phys
        
        # Check if DN is in correct format
        if not dn or '/phys-[' not in dn:
            print(f"Invalid DN format for physical details: {dn}")
            return {'operSt': 'down', 'operStQual': '', 'lastLinkStChg': ''}
            
        # Ensure DN doesn't already have /phys at the end
        if dn.endswith('/phys'):
            phys_dn = dn
        else:
            phys_dn = f"{dn}/phys"
        
        url = f"{self.apic_url}/api/mo/{phys_dn}.json"
        try:
            response = self._make_request('GET', url)
            response.raise_for_status()
            data = response.json()
            
            if data.get('imdata'):
                phys_if = data['imdata'][0].get('ethpmPhysIf', {}).get('attributes', {})
                return {
                    'operSt': phys_if.get('operSt', 'down').lower(),
                    'operStQual': phys_if.get('operStQual', ''),
                    'lastLinkStChg': phys_if.get('lastLinkStChg', '')
                }
            else:
                print(f"No physical interface data found for {phys_dn}")
        except Exception as e:
            print(f"Error getting physical details for {phys_dn}: {str(e)}")
        
        return {'operSt': 'down', 'operStQual': '', 'lastLinkStChg': ''}

    def _fetch_interface_details(self, dn, transceiver_lookup):
        """Fetch all details for a single interface (for parallel processing)"""
        try:
            # Get physical interface details
            phys_details = self.get_physical_details(dn)
            
            # Get transceiver info for this interface
            transceiver = transceiver_lookup.get(dn, {})
            
            # Get Ethernet statistics for this interface
            etherstats = self.get_ethernet_statistics(dn)
            
            return {
                'dn': dn,
                'phys_details': phys_details,
                'transceiver': transceiver,
                'etherstats': etherstats
            }
        except Exception as e:
            self.debug_print(f"Error fetching details for {dn}: {str(e)}")
            return {
                'dn': dn,
                'phys_details': {'operSt': 'down', 'operStQual': '', 'lastLinkStChg': ''},
                'transceiver': {},
                'etherstats': self._get_default_etherstats()
            }


    def to_xml(self):
        """Convert interface information to XML format"""
        root = ET.Element("InterfaceInformation")
        root.set("timestamp", datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
        
        for interface in self.interfaces:
            if_elem = ET.SubElement(root, "Interface")
            ET.SubElement(if_elem, "Node").text = interface.get('node', '')
            ET.SubElement(if_elem, "Name").text = interface.get('interface', '')
            ET.SubElement(if_elem, "Description").text = interface.get('description', '')
            ET.SubElement(if_elem, "AdminState").text = interface.get('adminSt', '')
            ET.SubElement(if_elem, "SwitchingState").text = interface.get('switchingSt', '')
            ET.SubElement(if_elem, "OperState").text = interface.get('operSt', '')
            ET.SubElement(if_elem, "OperReason").text = interface.get('operStQual', '')
            ET.SubElement(if_elem, "LastStateChange").text = interface.get('lastLinkStChg', '')
            ET.SubElement(if_elem, "Speed").text = interface.get('speed', '')
            ET.SubElement(if_elem, "Layer").text = interface.get('layer', '')
            ET.SubElement(if_elem, "Usage").text = interface.get('usage', '')
            ET.SubElement(if_elem, "MTU").text = interface.get('mtu', '')
            ET.SubElement(if_elem, "Hostname").text = interface.get('hostname', '')
            ET.SubElement(if_elem, "NodeModel").text = interface.get('nodeModel', '')
            ET.SubElement(if_elem, "NodeRole").text = interface.get('nodeRole', '')
            
            # Add transceiver information
            transceiver = ET.SubElement(if_elem, "Transceiver")
            ET.SubElement(transceiver, "Type").text = interface.get('transceiver', {}).get('type', '')
            ET.SubElement(transceiver, "Serial").text = interface.get('transceiver', {}).get('serial', '')
            ET.SubElement(transceiver, "Vendor").text = interface.get('transceiver', {}).get('vendor', '')
            
            # Add Ethernet statistics information
            etherstats = ET.SubElement(if_elem, "EthernetStatistics")
            etherstats_data = interface.get('etherstats', {})
            ET.SubElement(etherstats, "BroadcastPkts").text = etherstats_data.get('broadcastPkts', '0')
            ET.SubElement(etherstats, "ClearTs").text = etherstats_data.get('clearTs', 'never')
            ET.SubElement(etherstats, "MulticastPkts").text = etherstats_data.get('multicastPkts', '0')
            ET.SubElement(etherstats, "CRCAlignErrors").text = etherstats_data.get('cRCAlignErrors', '0')
            ET.SubElement(etherstats, "DropEvents").text = etherstats_data.get('dropEvents', '0')
            ET.SubElement(etherstats, "Fragments").text = etherstats_data.get('fragments', '0')
            ET.SubElement(etherstats, "Jabbers").text = etherstats_data.get('jabbers', '0')
            ET.SubElement(etherstats, "TotalPkts").text = etherstats_data.get('pkts', '0')
            ET.SubElement(etherstats, "OversizePkts").text = etherstats_data.get('oversizePkts', '0')
            ET.SubElement(etherstats, "RXNoErrors").text = etherstats_data.get('rXNoErrors', '0')
            ET.SubElement(etherstats, "RXGiantPkts").text = etherstats_data.get('rxGiantPkts', '0')
            ET.SubElement(etherstats, "RXOversizePkts").text = etherstats_data.get('rxOversizePkts', '0')
            ET.SubElement(etherstats, "TXNoErrors").text = etherstats_data.get('tXNoErrors', '0')
            ET.SubElement(etherstats, "TXGiantPkts").text = etherstats_data.get('txGiantPkts', '0')
            ET.SubElement(etherstats, "TXOversizePkts").text = etherstats_data.get('txOversizePkts', '0')
            ET.SubElement(etherstats, "UndersizePkts").text = etherstats_data.get('undersizePkts', '0')
        
        return root

    def to_excel(self, filename):
        """Export interface information to Excel format"""
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library is not installed. Please install it using: pip install openpyxl")
            return False
            
        try:
            # Create a new workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Interface Information"
            
            # Define headers
            headers = [
                "Node", "Interface", "Description", "Admin State", "Switching State", 
                "Oper State", "Oper Reason", "Speed", "Layer", "Usage", "MTU", 
                "Hostname", "Node Model", "Node Role", "Transceiver Type", 
                "Transceiver Serial", "Transceiver Vendor", "Broadcast Pkts", 
                "Multicast Pkts", "CRC Align Errors", "Drop Events", "Fragments", 
                "Jabbers", "Total Pkts", "Oversize Pkts", "RX No Errors", 
                "RX Giant Pkts", "RX Oversize Pkts", "TX No Errors", 
                "TX Giant Pkts", "TX Oversize Pkts", "Undersize Pkts"
            ]
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Add data rows
            for row, interface in enumerate(self.interfaces, 2):
                etherstats = interface.get('etherstats', {})
                transceiver = interface.get('transceiver', {})
                
                row_data = [
                    interface.get('node', ''),
                    interface.get('interface', ''),
                    interface.get('description', ''),
                    interface.get('adminSt', ''),
                    interface.get('switchingSt', ''),
                    interface.get('operSt', ''),
                    interface.get('operStQual', ''),
                    interface.get('speed', ''),
                    interface.get('layer', ''),
                    interface.get('usage', ''),
                    interface.get('mtu', ''),
                    interface.get('hostname', ''),
                    interface.get('nodeModel', ''),
                    interface.get('nodeRole', ''),
                    transceiver.get('type', ''),
                    transceiver.get('serial', ''),
                    transceiver.get('vendor', ''),
                    etherstats.get('broadcastPkts', '0'),
                    etherstats.get('multicastPkts', '0'),
                    etherstats.get('cRCAlignErrors', '0'),
                    etherstats.get('dropEvents', '0'),
                    etherstats.get('fragments', '0'),
                    etherstats.get('jabbers', '0'),
                    etherstats.get('pkts', '0'),
                    etherstats.get('oversizePkts', '0'),
                    etherstats.get('rXNoErrors', '0'),
                    etherstats.get('rxGiantPkts', '0'),
                    etherstats.get('rxOversizePkts', '0'),
                    etherstats.get('tXNoErrors', '0'),
                    etherstats.get('txGiantPkts', '0'),
                    etherstats.get('txOversizePkts', '0'),
                    etherstats.get('undersizePkts', '0')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    
                    # Highlight down interfaces
                    if col == 6 and value == 'down':  # Oper State column
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                # Check header length
                max_length = max(max_length, len(str(headers[col-1])))
                
                # Check data lengths
                for row in range(2, len(self.interfaces) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws.title = "Summary"
            
            # Summary statistics
            total_interfaces = len(self.interfaces)
            up_interfaces = sum(1 for iface in self.interfaces if iface.get('operSt') == 'up')
            down_interfaces = total_interfaces - up_interfaces
            
            summary_data = [
                ["Interface Summary", ""],
                ["Total Interfaces", total_interfaces],
                ["Up Interfaces", up_interfaces],
                ["Down Interfaces", down_interfaces],
                ["", ""],
                ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["", ""],
                ["Node Distribution", ""]
            ]
            
            # Count interfaces per node
            node_counts = {}
            for interface in self.interfaces:
                node = interface.get('node', 'Unknown')
                node_counts[node] = node_counts.get(node, 0) + 1
            
            for node, count in sorted(node_counts.items()):
                summary_data.append([node, count])
            
            # Add summary data to worksheet
            for row, (label, value) in enumerate(summary_data, 1):
                summary_ws.cell(row=row, column=1, value=label)
                summary_ws.cell(row=row, column=2, value=value)
                
                if row == 1:  # Header row
                    summary_ws.cell(row=row, column=1).font = header_font
                    summary_ws.cell(row=row, column=1).fill = header_fill
            
            # Auto-adjust summary column widths
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 15
            
            # Save the workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"{filename}_{timestamp}.xlsx"
            wb.save(excel_filename)
            
            print(f"\nExcel file saved as: {excel_filename}")
            print(f"Total interfaces exported: {total_interfaces}")
            print(f"Up interfaces: {up_interfaces}")
            print(f"Down interfaces: {down_interfaces}")
            
            return True
            
        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            return False

    def save_interface_info(self, filename, output_format='xml'):
        """Save interface and transceiver information to XML or Excel file"""
        interface_data = self.get_interface_info()
        transceiver_data = self.get_transceiver_info()
        fault_data = self.get_interface_faults()
        
        if not interface_data:
            print("\nNo interface data received from API")
            return

        # Create lookup for transceivers by DN
        transceiver_lookup = {}
        for item in transceiver_data.get('imdata', []):
            if 'ethpmFcot' in item:
                fcot = item['ethpmFcot']['attributes']
                dn = fcot.get('dn', '')
                # Get the interface DN part (remove /phys/fcot)
                interface_dn = '/'.join(dn.split('/')[:-2])
                transceiver_lookup[interface_dn] = fcot

        # Create lookup for faults by DN
        fault_lookup = {}
        if fault_data:
            for item in fault_data.get('imdata', []):
                if 'faultInst' in item:
                    fault = item['faultInst']['attributes']
                    dn = fault.get('dn', '')
                    # Extract interface DN from fault DN
                    if '/phys-[' in dn:
                        interface_dn = dn.split('/fault-')[0]
                        if interface_dn not in fault_lookup:
                            fault_lookup[interface_dn] = []
                        fault_lookup[interface_dn].append(fault)

        # Get current timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Ensure data directory exists
        data_dir = os.path.join('data')
        os.makedirs(data_dir, exist_ok=True)
        
        xml_filename = os.path.join(data_dir, f"{filename}_{timestamp}.xml")
        
        # Initialize table headers and data
        headers = ["Node", "Interface", "Description", "Admin State", "Switching State", "Oper State", "Oper Reason", "Speed", "Layer", "Usage", "MTU", "Hostname", "Node Model", "Node Role", "Type", "Serial", "Vendor", "Broadcast Pkts", "Multicast Pkts", "CRC Align Errors", "Drop Events", "Fragments", "Jabbers", "Total Pkts", "Oversize Pkts", "RX No Errors", "RX Giant Pkts", "RX Oversize Pkts", "TX No Errors", "TX Giant Pkts", "TX Oversize Pkts", "Undersize Pkts"]
        table_data = []
        
        if interface_data:
            print(f"\nFound {len(interface_data)} interface entries")
            print("Collecting interface details, transceiver info, and Ethernet statistics...")
            print("Using parallel processing for faster execution...")
            
            # Prepare list of DNs to process
            dns_to_process = []
            for item in interface_data:
                dn = item.get('dn', '')
                pod, node, interface = self.parse_dn(dn)
                # Only process physical ethernet interfaces
                if interface and interface.startswith('eth'):
                    dns_to_process.append((dn, item))
            
            print(f"Processing {len(dns_to_process)} ethernet interfaces in parallel...")
            
            # Process interfaces in parallel using ThreadPoolExecutor
            details_map = {}
            with ThreadPoolExecutor(max_workers=20) as executor:
                # Submit all tasks
                future_to_dn = {
                    executor.submit(self._fetch_interface_details, dn, transceiver_lookup): dn 
                    for dn, _ in dns_to_process
                }
                
                # Collect results as they complete
                completed = 0
                for future in as_completed(future_to_dn):
                    result = future.result()
                    details_map[result['dn']] = result
                    completed += 1
                    
                    # Show progress every 100 interfaces
                    if completed % 100 == 0:
                        print(f"  Processed {completed}/{len(dns_to_process)} interfaces...")
            
            print(f"  Completed processing all {len(dns_to_process)} interfaces!")
            
            # Now build the interface info and table data
            for dn, item in dns_to_process:
                details = details_map.get(dn)
                if not details:
                    continue
                
                phys_details = details['phys_details']
                transceiver = details['transceiver']
                etherstats = details['etherstats']
                
                # Parse DN for pod/node/interface
                pod, node, interface = self.parse_dn(dn)
                        
                # Get admin and switching states
                admin_state = item.get('adminSt', 'down').lower()
                switching_state = item.get('switchingSt', 'disabled').lower()
                
                # Ensure admin state is only 'up' or 'down'
                admin_state = 'up' if admin_state == 'up' else 'down'

                # Add additional state info if operationally down
                if phys_details['operSt'] == 'down':
                    self.debug_print(f"\nInterface {interface} on Node-{node} is operationally down:")
                    self.debug_print(f"  Operational State: {phys_details['operSt']}")
                    self.debug_print(f"  Reason: {phys_details['operStQual']}")
                    self.debug_print(f"  Last State Change: {phys_details['lastLinkStChg']}")
                    
                # Create interface info
                interface_info = {
                    'pod': f"Pod-{pod}" if pod else "N/A",
                    'node': f"Node-{node}" if node else "N/A",
                    'interface': interface if interface else "N/A",
                    'description': item.get('descr', 'N/A'),
                    'adminSt': admin_state,
                    'switchingSt': switching_state,
                    'operSt': phys_details['operSt'],
                    'operStQual': phys_details['operStQual'],
                    'lastLinkStChg': phys_details['lastLinkStChg'],
                    'speed': item.get('speed', 'N/A'),
                    'layer': item.get('layer', 'N/A'),
                    'usage': item.get('usage', 'N/A'),
                    'mtu': item.get('mtu', 'N/A'),
                    'hostname': item.get('hostname', 'N/A'),
                    'nodeModel': item.get('nodeModel', 'N/A'),
                    'nodeRole': item.get('nodeRole', 'N/A'),
                    'transceiver': {
                        'type': transceiver.get('typeName', ''),
                        'serial': transceiver.get('guiSN', ''),
                        'vendor': transceiver.get('guiName', '')
                    },
                    'etherstats': etherstats
                }
                self.interfaces.append(interface_info)
                
                # Add to table data
                table_data.append([
                    f"Node-{node}" if node else "N/A",
                    interface if interface else "N/A",
                    item.get('descr', 'N/A'),
                    admin_state,
                    switching_state,
                    phys_details['operSt'],
                    phys_details['operStQual'],
                    item.get('speed', 'N/A'),
                    item.get('layer', 'N/A'),
                    item.get('usage', 'N/A'),
                    item.get('mtu', 'N/A'),
                    item.get('hostname', 'N/A'),
                    item.get('nodeModel', 'N/A'),
                    item.get('nodeRole', 'N/A'),
                    transceiver.get('typeName', ''),
                    transceiver.get('guiSN', ''),
                    transceiver.get('guiName', ''),
                    etherstats.get('broadcastPkts', '0'),
                    etherstats.get('multicastPkts', '0'),
                    etherstats.get('cRCAlignErrors', '0'),
                    etherstats.get('dropEvents', '0'),
                    etherstats.get('fragments', '0'),
                    etherstats.get('jabbers', '0'),
                    etherstats.get('pkts', '0'),
                    etherstats.get('oversizePkts', '0'),
                    etherstats.get('rXNoErrors', '0'),
                    etherstats.get('rxGiantPkts', '0'),
                    etherstats.get('rxOversizePkts', '0'),
                    etherstats.get('tXNoErrors', '0'),
                    etherstats.get('txGiantPkts', '0'),
                    etherstats.get('txOversizePkts', '0'),
                    etherstats.get('undersizePkts', '0')
                ])
        
        # Sort table data by node and interface
        table_data.sort(key=lambda x: (x[0], x[1]))
        
        # Save to file based on output format
        if output_format.lower() == 'excel':
            # Save to Excel file
            if self.to_excel(filename):
                # Display summary table
                print("\nInterface Summary:")
                print(tabulate(table_data, headers=headers, tablefmt="pretty", numalign="left", stralign="left"))
            else:
                print("Failed to save Excel file")
        else:
            # Save to XML file with pretty printing (default)
            try:
                xml_str = ET.tostring(self.to_xml(), encoding='unicode')
                pretty_xml = xml.dom.minidom.parseString(xml_str).toprettyxml()
                
                with open(xml_filename, 'w') as f:
                    f.write(pretty_xml)
                
                print(f"\nSaved interface information to {xml_filename}")
                
                # Display summary table
                print("\nInterface Summary:")
                print(tabulate(table_data, headers=headers, tablefmt="pretty", numalign="left", stralign="left"))
                
            except Exception as e:
                print(f"Failed to save XML file: {str(e)}")

def get_credentials():
    """Prompt for APIC credentials or load from environment"""
    
    # Try to get from environment variables first
    env_ip = os.getenv('APIC_IP')
    env_user = os.getenv('APIC_USERNAME')
    env_password = os.getenv('APIC_PASSWORD')
    
    if env_ip and env_user and env_password:
        print("Using credentials from environment variables")
        # Ensure URL starts with https://
        if not env_ip.startswith("https://"):
            env_ip = f"https://{env_ip}"
        return env_ip, "DefaultAuth", env_user, env_password

    print("\nEnter APIC connection details:")
    while True:
        apic = input("APIC IP/hostname [https://172.24.207.2]: ").strip()
        if not apic:
            apic = "https://172.24.207.2"
        
        # Ensure URL starts with https://
        if not apic.startswith("https://"):
            apic = f"https://{apic}"

        # Create temporary instance to get auth domains
        temp_aci = ACIInterfaceInfo(apic, "", "")
        print("\nRetrieving authentication domains...")
        domains = temp_aci.list_auth_domains()
        
        if domains:
            print("\nAvailable authentication domains:")
            for i, domain in enumerate(domains, 1):
                print(f"{i}. {domain['name']} ({domain['type']})")
            
            while True:
                try:
                    choice = int(input("\nSelect authentication domain (1-{}): ".format(len(domains))))
                    if 1 <= choice <= len(domains):
                        selected_domain = domains[choice-1]['name']
                        break
                    else:
                        print("Invalid choice. Please try again.")
                except ValueError:
                    print("Please enter a number.")
        else:
            print("Could not retrieve authentication domains. Using default authentication.")
            selected_domain = "DefaultAuth"

        max_attempts = 3
        for attempt in range(max_attempts):
            username = input("Username [admin]: ").strip()
            if not username:
                username = "admin"
                
            password = getpass.getpass("Password: ")
            if not password:
                print("Password cannot be empty. Please try again.")
                continue

            # Create a test instance to verify credentials
            test_aci = ACIInterfaceInfo(apic, username, password)
            if test_aci.login(selected_domain):
                return apic, selected_domain, username, password
            
            attempts_left = max_attempts - attempt - 1
            if attempts_left > 0:
                print(f"\nLogin failed. {attempts_left} attempts remaining. Please try again.")
            else:
                print("\nMaximum login attempts exceeded. Exiting.")
                sys.exit(1)

def main():
    parser = argparse.ArgumentParser(description='Get ACI interface information')
    parser.add_argument('-a', '--apic', help='APIC IP/hostname (e.g., https://172.24.207.2)')
    parser.add_argument('-u', '--username', help='APIC username')
    parser.add_argument('-p', '--password', help='APIC password')
    parser.add_argument('-f', '--filename', default='interface_info', help='Base filename for output (default: interface_info)')
    parser.add_argument('-d', '--debug', action='store_true', help='Enable debug output')
    parser.add_argument('-o', '--output', choices=['xml', 'excel'], default='xml', help='Output format: xml or excel (default: xml)')
    args = parser.parse_args()

    # Get credentials either from arguments or prompt
    if args.apic and args.username and args.password:
        apic_url = args.apic
        username = args.username
        password = args.password
        selected_domain = "DefaultAuth"  # Use default for command line args
    else:
        apic_url, selected_domain, username, password = get_credentials()

    # Initialize ACI interface info
    aci = ACIInterfaceInfo(apic_url, username, password)
    
    # Enable debug if requested
    if args.debug:
        aci.set_debug(True)
    
    # Login to APIC with selected domain
    if not aci.login(selected_domain):
        print("Failed to login to APIC")
        return
    
    # Save interface info
    aci.save_interface_info(args.filename, args.output)

if __name__ == "__main__":
    main()
